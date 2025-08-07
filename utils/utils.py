import re
from loguru import logger as _logger
import pandas as pd
import os
import openpyxl


def calculate_f1_score_from_csv(file_path: str) -> float:
    """
    读取指定的CSV文件，并根据TP, FP, FN列计算F1-score。

    Args:
        file_path (str): CSV文件的路径。

    Returns:
        float: 计算出的F1-score。如果TP+FP或TP+FN为0，则返回0.0。
    """
    try:
        # 使用pandas读取CSV文件
        df = pd.read_csv(file_path)

        # 确保必需的列存在
        required_columns = {'TP', 'FP', 'FN'}
        if not required_columns.issubset(df.columns):
            raise ValueError(f"CSV文件中缺少必需的列，需要包含: {', '.join(required_columns)}")

        # 计算所有行的TP, FP, FN的总和
        total_tp = df['TP'].sum()
        total_fp = df['FP'].sum()
        total_fn = df['FN'].sum()

        print(f"总 True Positives (TP): {total_tp}")
        print(f"总 False Positives (FP): {total_fp}")
        print(f"总 False Negatives (FN): {total_fn}")

        # 计算精确率 (Precision)
        # Precision = TP / (TP + FP)
        if total_tp + total_fp == 0:
            precision = 0.0
        else:
            precision = total_tp / (total_tp + total_fp)

        # 计算召回率 (Recall)
        # Recall = TP / (TP + FN)
        if total_tp + total_fn == 0:
            recall = 0.0
        else:
            recall = total_tp / (total_tp + total_fn)

        print(f"精确率 (Precision): {precision:.4f}")
        print(f"召回率 (Recall): {recall:.4f}")

        # 计算F1-score
        # F1-score = 2 * (Precision * Recall) / (Precision + Recall)
        if precision + recall == 0:
            f1_score = 0.0
        else:
            f1_score = 2 * (precision * recall) / (precision + recall)

        return f1_score

    except FileNotFoundError:
        print(f"错误：文件 '{file_path}' 未找到。")
        return 0.0
    except Exception as e:
        print(f"处理文件时发生错误: {e}")
        return 0.0


def get_logger():
    """
    自定义logger
    :return:
    """
    _logger.add("./logs/run.log", rotation="500 MB", encoding="utf-8")
    return _logger


def md_to_dict_str(text: str) -> str:
    """
    从Markdown文本中提取字典字符串
    :param text: Markdown文本
    :return: 提取的字典字符串，如果未找到则返回空字符串
    """
    pattern = r'(\{.*?\})'  # 匹配花括号内的内容
    match = re.search(pattern, text, re.DOTALL)  # 查找匹配项
    return match.group(1) if match else ''  # 返回匹配的字典字符串，否则返回空字符串


logger = get_logger()


def create_multisheet_excel_report(csv_list: list[str], output_excel_path: str):
    """
    读取指定目录下的所有CSV文件，将每个文件写入一个Excel工作表，
    并创建一个包含摘要信息的额外工作表。

    Args:
        csv_directory (str): 包含CSV文件的目录路径。
        output_excel_path (str): 输出的Excel文件路径。
    """
    # 用于存储摘要信息
    summary_data = []

    # 2. 使用 pd.ExcelWriter 作为上下文管理器，确保文件被正确保存和关闭
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:

        # 3. 遍历所有CSV文件
        for csv_file in csv_list:
            try:
                # 读取CSV文件到pandas DataFrame
                df = pd.read_csv(csv_file)

                # 从文件名创建一个简洁的sheet名称 (例如 'sales_data_2023.csv' -> 'sales_data_2023')
                sheet_name = csv_file.split('/')[-2] + '#' + csv_file.split('/')[-1].split('.')[0]
                # 确保sheet名长度不超过Excel限制（31个字符）
                sheet_name = sheet_name[:31]

                # 将DataFrame写入Excel的一个新sheet
                # index=False 参数可以防止pandas将DataFrame的索引写入Excel
                print(f"正在将 '{csv_file}' 写入 sheet '{sheet_name}'...")
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # 收集摘要信息（例如：文件名和行数）
                summary_data.append({
                    '来源文件': csv_file,
                    '工作表名称': sheet_name,
                    '行数': len(df),
                    '列数': len(df.columns)
                })

            except Exception as e:
                print(f"处理文件 '{csv_file}' 时出错: {e}")
                # 记录错误信息到摘要中
                summary_data.append({
                    '来源文件': csv_file,
                    '工作表名称': 'ERROR',
                    '行数': 0,
                    '列数': 0
                })

        # 4. 创建并写入摘要信息sheet
        print("正在创建摘要信息 sheet...")
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

    print(f"\n处理完成！报告已保存至 '{output_excel_path}'")


# --- 新建函数 ---
def add_f1_score_formulas_to_report(excel_path: str):
    """
    打开一个现有的Excel报告，并在'Summary'表中添加F1-score计算公式。

    Args:
        excel_path (str): Excel文件的路径。
    """
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        print(f"错误：无法添加公式，因为文件 '{excel_path}' 未找到。")
        return

    if 'Summary' not in workbook.sheetnames:
        print("错误：在工作簿中未找到 'Summary' 工作表。")
        return

    summary_sheet = workbook['Summary']

    # 1. 找到“工作表名称”所在的列索引 (从1开始)
    try:
        header_cells = summary_sheet[1]
        sheet_name_col_idx = [c.value for c in header_cells].index('工作表名称') + 1
    except ValueError:
        print("错误：在 'Summary' 工作表中未找到 '工作表名称' 列。")
        return

    # 2. 在 'Summary' 表的表头添加新列
    headers_to_add = ['Total TP', 'Total FP', 'Total FN', 'Precision', 'Recall', 'F1-Score']
    start_col = summary_sheet.max_column + 1
    for i, header in enumerate(headers_to_add):
        summary_sheet.cell(row=1, column=start_col + i, value=header)

    # 3. 遍历 'Summary' 表的每一行 (从第二行开始)，并写入公式
    for row in range(2, summary_sheet.max_row + 1):
        sheet_name = summary_sheet.cell(row=row, column=sheet_name_col_idx).value

        # 如果工作表名称无效或为错误，则跳过
        if not sheet_name or sheet_name == 'ERROR':
            continue

        # 因为工作表名称中可能包含特殊字符'#'，需要用单引号包裹
        safe_sheet_name = f"'{sheet_name}'"

        # 根据您的数据格式，TP在B列, FP在C列, FN在D列
        tp_formula = f"=SUM({safe_sheet_name}!B:B)"
        fp_formula = f"=SUM({safe_sheet_name}!C:C)"
        fn_formula = f"=SUM({safe_sheet_name}!D:D)"

        # 写入TP, FP, FN的总和公式
        tp_cell = summary_sheet.cell(row=row, column=start_col)
        fp_cell = summary_sheet.cell(row=row, column=start_col + 1)
        fn_cell = summary_sheet.cell(row=row, column=start_col + 2)

        tp_cell.value = tp_formula
        fp_cell.value = fp_formula
        fn_cell.value = fn_formula

        # 获取这些单元格的地址 (例如 'E2', 'F2', 'G2')
        tp_addr = tp_cell.coordinate
        fp_addr = fp_cell.coordinate
        fn_addr = fn_cell.coordinate

        # 创建 Precision, Recall, 和 F1-score 的公式
        # 使用 IFERROR 处理除以零的情况
        precision_formula = f"=IFERROR({tp_addr}/({tp_addr}+{fp_addr}), 0)"
        recall_formula = f"=IFERROR({tp_addr}/({tp_addr}+{fn_addr}), 0)"

        # 写入 Precision 和 Recall 公式，并获取它们的地址
        precision_cell = summary_sheet.cell(row=row, column=start_col + 3)
        recall_cell = summary_sheet.cell(row=row, column=start_col + 4)
        precision_cell.value = precision_formula
        recall_cell.value = recall_formula

        precision_addr = precision_cell.coordinate
        recall_addr = recall_cell.coordinate

        # 最后，创建 F1-score 公式
        f1_formula = f"=IFERROR(2*({precision_addr}*{recall_addr})/({precision_addr}+{recall_addr}), 0)"
        summary_sheet.cell(row=row, column=start_col + 5).value = f1_formula

    # 4. 保存修改后的工作簿
    workbook.save(excel_path)
    print(f"成功将F1-score计算公式添加到 '{excel_path}'")


# --- 新建函数 ---
def add_f1_score_formulas_to_report(excel_path: str):
    """
    打开一个现有的Excel报告，并在'Summary'表中添加F1-score计算公式。

    Args:
        excel_path (str): Excel文件的路径。
    """
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        print(f"错误：无法添加公式，因为文件 '{excel_path}' 未找到。")
        return

    if 'Summary' not in workbook.sheetnames:
        print("错误：在工作簿中未找到 'Summary' 工作表。")
        return

    summary_sheet = workbook['Summary']

    # 1. 找到“工作表名称”所在的列索引 (从1开始)
    try:
        header_cells = summary_sheet[1]
        sheet_name_col_idx = [c.value for c in header_cells].index('工作表名称') + 1
    except ValueError:
        print("错误：在 'Summary' 工作表中未找到 '工作表名称' 列。")
        return

    # 2. 在 'Summary' 表的表头添加新列
    headers_to_add = ['Total TP', 'Total FP', 'Total FN', 'Precision', 'Recall', 'F1-Score']
    start_col = summary_sheet.max_column + 1
    for i, header in enumerate(headers_to_add):
        summary_sheet.cell(row=1, column=start_col + i, value=header)

    # 3. 遍历 'Summary' 表的每一行 (从第二行开始)，并写入公式
    for row in range(2, summary_sheet.max_row + 1):
        sheet_name = summary_sheet.cell(row=row, column=sheet_name_col_idx).value

        # 如果工作表名称无效或为错误，则跳过
        if not sheet_name or sheet_name == 'ERROR':
            continue

        # 因为工作表名称中可能包含特殊字符'#'，需要用单引号包裹
        safe_sheet_name = f"'{sheet_name}'"

        # 根据您的数据格式，TP在B列, FP在C列, FN在D列
        tp_formula = f"=SUM({safe_sheet_name}!B:B)"
        fp_formula = f"=SUM({safe_sheet_name}!C:C)"
        fn_formula = f"=SUM({safe_sheet_name}!D:D)"

        # 写入TP, FP, FN的总和公式
        tp_cell = summary_sheet.cell(row=row, column=start_col)
        fp_cell = summary_sheet.cell(row=row, column=start_col + 1)
        fn_cell = summary_sheet.cell(row=row, column=start_col + 2)

        tp_cell.value = tp_formula
        fp_cell.value = fp_formula
        fn_cell.value = fn_formula

        # 获取这些单元格的地址 (例如 'E2', 'F2', 'G2')
        tp_addr = tp_cell.coordinate
        fp_addr = fp_cell.coordinate
        fn_addr = fn_cell.coordinate

        # 创建 Precision, Recall, 和 F1-score 的公式
        # 使用 IFERROR 处理除以零的情况
        precision_formula = f"=IFERROR({tp_addr}/({tp_addr}+{fp_addr}), 0)"
        recall_formula = f"=IFERROR({tp_addr}/({tp_addr}+{fn_addr}), 0)"

        # 写入 Precision 和 Recall 公式，并获取它们的地址
        precision_cell = summary_sheet.cell(row=row, column=start_col + 3)
        recall_cell = summary_sheet.cell(row=row, column=start_col + 4)
        precision_cell.value = precision_formula
        recall_cell.value = recall_formula

        precision_addr = precision_cell.coordinate
        recall_addr = recall_cell.coordinate

        # 最后，创建 F1-score 公式
        f1_formula = f"=IFERROR(2*({precision_addr}*{recall_addr})/({precision_addr}+{recall_addr}), 0)"
        summary_sheet.cell(row=row, column=start_col + 5).value = f1_formula

    # 4. 保存修改后的工作簿
    workbook.save(excel_path)
    print(f"成功将F1-score计算公式添加到 '{excel_path}'")


if __name__ == '__main__':
    # 定义输入和输出路径
    csv_files_to_process = [
        './data/Qwen3-14B-bid/f1_score.csv',
        './data/gpt-oss-20b-bid/f1_score.csv',
    ]
    report_path = "./data/f1_score_report.xlsx"

    # 第一步：创建包含原始数据的多工作表Excel报告
    create_multisheet_excel_report(
        csv_list=csv_files_to_process,
        output_excel_path=report_path
    )

    # 第二步：向生成的报告中添加F1-score计算公式
    add_f1_score_formulas_to_report(excel_path=report_path)
